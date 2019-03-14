import os, sys, xlwt
import xlrd
# from xlutils.copy import copy
import openpyxl
import re

def file_path(file_dir):
    #读取指定目录下的所有.java文件并存入列表中
    file = []

    for root, dirs, files in os.walk(file_dir):
        # print(root, end = ' ') #当前目录路径
        # print(dirs, end = ' ') #当前路径下的所有子目录
        # print(files)           #当前目录下的所有非目录子文件
    #for root, files in os.walk(file_dir):   
        if files:
            for file1 in files:
                if file1.endswith('.java'):
                    file1 = root + '\\' + file1
                    file.append(file1)

    return file


def getjava_test(java_files, test_files):
    java_test = []

    for java_file in java_files:
        i1 = len(java_test)
        java_name = java_file.split('\\')[-1]
        java_name = java_name.split('.')[0]

        for test_file in test_files:
            test_name = test_file.split('\\')[-1]
            test_name = test_name.split('.')[0]
            test_name = re.sub('Test', '', test_name)

            if java_name == test_name:
                java_test1 = {'file': java_file, 'test': test_file}
                java_test.append(java_test1)
                test_files.remove(test_file)
                break
        
        i2 = len(java_test)
        if(i1 == i2):
            java_test1 = {'file': java_file, 'test': 'null'}
            java_test.append(java_test1)
    
    if(len(test_files)):
        for test_file in test_files:
            java_test1 = {'file': 'null', 'test': test_file}
            java_test.append(java_test1)
    
    return java_test


def write_excel(path, name_list):
    name = []
    #创建工作薄
    # workbook = xlwt.Workbook(encoding = 'utf-8')
    workbook = openpyxl.Workbook()
    # workbook = workbook1.active
    #创建sheet
    data_sheet = workbook.create_sheet('demo')

    for i in range(len(name_list)):
        # data_sheet.write(i, 0, name_list[i]['file'])
        # data_sheet.write(i, 1, name_list[i]['test'])
        data_sheet.cell(row=i+1, column=1, value=name_list[i]['file'])
        data_sheet.cell(row=i+1, column=2, value=name_list[i]['test'])
        
        sheet_name = str(i + 1)
        sheet = workbook.create_sheet(sheet_name)
        sheet.cell(row=1, column=1, value='commit_id')
        sheet.cell(row=1, column=2, value=name_list[i]['file'].split('\\')[-1])
        sheet.cell(row=1, column=3, value=name_list[i]['test'].split('\\')[-1])
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 100
        sheet.column_dimensions["C"].width = 100

        if name_list[i]['file'] == 'null':
            name.append(re.sub('Test', '', name_list[i]['test'].split('\\')[-1].rstrip('.java')))
        else:
            name.append(name_list[i]['file'].split('\\')[-1].rstrip('.java'))

    #设置单元格宽度
    data_sheet.column_dimensions["A"].width = 120
    data_sheet.column_dimensions["B"].width = 120
    workbook.save(path)
    
    return name


def change_sheet(name, excel_path, input_dir): 
    wb = openpyxl.load_workbook(excel_path)

    for root, dirs, files in os.walk(input_dir):
        if files:
            for file1 in files:
                name1 =  file1.split('.')[0]
                name2 = re.sub('Test', '', name1)
                
                if name2 in name:
                    i = name.index(name2) + 1
                    sheet = wb.get_sheet_by_name(str(i))
                    # k = len(sheet.rows)
                    k = sheet.max_row
                    
                    root = root.split('\\')[-1]
                    if sheet.cell(k, 1).value != root:        
                        k = k + 1
                        sheet.cell(row=k, column=1, value=root)
                    if 'Test' in name1:
                        sheet.cell(row=k, column=3, value='1')
                    else:
                        sheet.cell(row=k, column=2, value='1')
                    sheet.column_dimensions["A"].width = 90
                    sheet.column_dimensions["B"].width = 50
                    sheet.column_dimensions["C"].width = 50

    wb.save(excel_path)


if __name__ == '__main__':
    print ("main文件夹名：",sys.argv[1])
    print ("test文件夹名：",sys.argv[2])
    main_dir = sys.argv[1]
    test_dir = sys.argv[2]
    # main_dir = 'E:\python_work\\test\\3_logstash-logback-encoder\src\main'
    # test_dir = 'E:\python_work\\test\\3_logstash-logback-encoder\src\\test'

    java_file = file_path(main_dir)
    test_file = file_path(test_dir)

    name_list = getjava_test(java_file, test_file)

    project_name = main_dir.split('\\')[-3]
    output_path = 'E:\output\\' + project_name
    isExists = os.path.exists(output_path)
    if not isExists:
      os.makedirs(output_path)

    path = output_path + '\\' + project_name + '.xlsx'
    name = write_excel(path, name_list)
    print(u'创建xlsx文件成功')

    # change_sheet(name, path, 'E:\output\\3_logstash-logback-encoder')
    change_sheet(name, path, output_path)
    print(u'填入成功')