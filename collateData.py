import openpyxl
import os
import shutil


def collectRepairCases(allcasenameFile, excelFile):
    #收集用例名称
    workbook = openpyxl.Workbook()
    data_sheet = workbook.create_sheet('result')
    i = 1
    with open(allcasenameFile) as file_object:   
        for line in file_object:
            i = i + 1
            data_sheet.cell(row=i, column=1, value=line)
    
    workbook.save(excelFile)


def collectResult(testrepairPath, excelFile):
    i = 1
    wb = openpyxl.load_workbook(excelFile)
    sheet = wb.get_sheet_by_name('result')
    for root, dirs, files in os.walk(testrepairPath):
        # print(root, end = ' ') #当前目录路径
        # print(dirs, end = ' ') #当前路径下的所有子目录
        # print(files)           #当前目录下的所有非目录子文件
        for file in files:
            if file == 'repair-test-case-log.txt':
                i = i + 1
                with open(os.path.join(root, file)) as file_object:
                    contents = file_object.read()
                    sheet.cell(row=i, column=4, value=contents)
            #    continue
    wb.save(excelFile)


def collectResult1(testrepairPath):
    i = 1
    workbook = openpyxl.Workbook()
    data_sheet = workbook.create_sheet('result')
    for root, dirs, files in os.walk(testrepairPath):
        for file in files:
            if file == 'repair-test-case-log.txt':
                i = i + 1
                with open(os.path.join(root, file)) as file_object:
                    data_sheet.cell(row=i, column=1, value=root)
                    contents = file_object.read()
                    data_sheet.cell(row=i, column=2, value=contents)
    workbook.save(testrepairPath + '\\data.xlsx')


def collectVariants(onetestrepairPath):
    i = 0
    for root, dirs, files in os.walk(onetestrepairPath):
        for file in files:
            if file.endswith('.java'):
                targetpath = onetestrepairPath + '\\variants-summary'
                if not os.path.exists(targetpath):
                    os.makedirs(onetestrepairPath + '\\variants-summary')
                shutil.copyfile(os.path.join(root, file), targetpath + '\\' + root.split('\\')[11] + '.java')
                i = i + 1


def collectOnetestrepairPath(testrepairPath):
    f = open(testrepairPath + '\\onetestrepairPath.txt', 'w+')
    for root, dirs, files in os.walk(testrepairPath):
        for file in files:
            if file == 'repair-test-case-log.txt':
                f.write(root + '\n')
    f.close()


def collectallVariants(OnetestrepairPathFile):
    with open(OnetestrepairPathFile) as file_object:
        for line in file_object:
            line = line.strip('\n')
            collectVariants(line)


def collectrightTest(allrepaircasesPath, repairSubjectsPath):
    #提取包含正确测试用例的文件
    with open(allrepaircasesPath) as file_object:
        for line in file_object:
            path = line.replace('\t', '\\')
            path1 = path.split('-vs-')[0]
            path2 = path.split('-vs-')[1]
            path2_1 = path2.split('\\')[1]
            path2_2 = path2_1.replace('.', '\\')
            path2_3 = path2_2.split('\\')[-1]
            path2_4 = path2_2.replace('\\' + path2_3, '.java')
            targetPath = repairSubjectsPath + '\\' + path1 + '\\test\\' + path2_4
            finalfileName = path1 + '\\test\\' + path2_4
            finalfileName = finalfileName.replace('\\', '_')
            finalfileName = 'C:\\Users\\Y_sun\\Desktop\\2019-1\\experiment-data\\data-summary\\correct-test\\' + finalfileName
            if not os.path.exists(finalfileName):
                shutil.copyfile(targetPath, finalfileName)


if __name__ == '__main__':
    # collectRepairCases('C:\\Users\\Y_sun\\Desktop\\2019-1\\experiment-data\\repair-subjects\\all-repair-cases.txt', 'C:\\Users\\Y_sun\\Desktop\\2019-1\\experiment-data\\data.xlsx')
    # collectResult('C:\\Users\\Y_sun\\Desktop\\2019-1\\experiment-data\\test-repair-author-0625-2min', 'C:\\Users\\Y_sun\\Desktop\\2019-1\\experiment-data\\data.xlsx')
    # collectResult1('C:\\Users\\Y_sun\\Desktop\\2019-1\\experiment-data\\test-repair-me-0702-2min')
    #  collectVariants('C:\\Users\\Y_sun\\Desktop\\2019-1\\experiment-data\\test-repair-me-0627-2min\\Gson\\gson-1.3-vs-gson-1.2\\com_google_gson_JsonEscapingVisitorTest_testStringArrayVisitationEscapingRequired')
    # collectOnetestrepairPath('C:\\Users\\Y_sun\\Desktop\\2019-1\\experiment-data\\test-repair-me-0624-2min')
    # collectallVariants('C:\\Users\\Y_sun\\Desktop\\2019-1\\experiment-data\\test-repair-me-0624-2min\\onetestrepairPath.txt')
    collectrightTest('C:\\Users\\Y_sun\\Desktop\\2019-1\\experiment-data\\repair-subjects\\all-repair-cases.txt', 'C:\\Users\\Y_sun\\Desktop\\2019-1\\experiment-data\\repair-subjects')